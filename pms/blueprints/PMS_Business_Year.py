from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from zipfile import BadZipFile

import mysql.connector
from flask import Blueprint, request, render_template, jsonify, session, url_for, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..db import create_connection
from ..services.progress import calc_progress, calc_progress_bulk

bp = Blueprint('business_year', __name__)


PROJECT_RAW_COLUMN_LABELS = {
    'ProjectID': '사업 ID',
    'ContractCode': '계약코드',
    'ProjectName': '사업명',
    'ProjectCost': '총사업비(VAT 포함)',
    'ProjectCost_NoVAT': '총사업비(VAT 제외)',
    'StartDate': '착수일',
    'EndDate': '준공일',
    'orderPlace': '발주처',
    'Manager': '담당자',
    'ContributionRate': '기술료율',
    'ProjectDetails': '사업개요',
    'AcademicResearchRate': '사전비용',
    'OperationalRate': '운영비용',
    'EquipmentRate': '공정비용',
    'ChangeProjectCost': '변경사업비',
    'outsourcingCheck': '외주 여부',
    'yearProject': '연차사업 여부',
    'LinkProjectCheck': '연계사업 여부',
    'referenceProjects': '참조사업',
    'safetyRate': '안전관리비율',
    'project_status': '사업상태',
    'BidPrice_NoVAT': '투찰가(VAT 제외)',
    'BidPrice': '투찰가(VAT 포함)',
    'D_Day': 'D-Day',
    'performanceRiview': '성과심사',
    'procurementType': '조달유형',
}

PROJECT_RAW_RELATED_TABLE_LABELS = {
    'businesschangehistory': '사업변경이력',
    'businessreceiptdetails': '사업비 수령내역',
    'clone_exmanager': '복제 외주인건비',
    'clone_expenserecords': '복제 경비내역',
    'clone_state': '복제 상태값',
    'examine_exmanager': '검토 외주인건비',
    'examine_expenserecords': '검토 경비내역',
    'examine_note': '검토 메모',
    'examine_outsourcing': '검토 외주내역',
    'exmanager': '외주인건비',
    'expenserecords': '경비내역',
    'external_labor_rates': '외부 노임단가',
    'meeting_files': '회의 파일',
    'outsourcing': '외주내역',
    'performanceevaluationfee': '성과심사비',
    'project_comment': '프로젝트 코멘트',
    'project_depbohal': '부서 보할',
    'project_engineers': '참여기술자',
    'project_risks': '프로젝트 리스크',
    'project_status_history': '상태 이력',
    'projectfiles': '프로젝트 파일',
    'quantity_log': '수량 로그',
    'state': '상태값',
    'taskassignment': '업무배정',
    'taskquantity': '수량산출',
    'transferdata': '이관데이터',
    'usemoney': '사용금액',
    'usemoney_log': '사용금액 로그',
}

PROJECT_RAW_REFERENCE_COLUMNS = [
    'referenceProject1',
    'referenceProject2',
    'referenceProject3',
    'referenceProject4',
    'referenceProject5',
]

PROJECT_RAW_EXCLUDED_COLUMNS = {
    'ProjectID',
    'safetyRate',
}

PROJECT_RAW_MONEY_COLUMNS = {
    'ProjectCost',
    'ProjectCost_NoVAT',
    'ChangeProjectCost',
    'BidPrice_NoVAT',
    'BidPrice',
}

PROJECT_RAW_BOOLEAN_COLUMNS = {
    'outsourcingCheck',
    'yearProject',
}

ASSET_MANAGEMENT_IMPORT_TABLE = 'asset_management_imports'
ASSET_MANAGEMENT_ROW_TABLE = 'asset_management_rows'
ASSET_MANAGEMENT_HISTORY_TABLE = 'asset_management_history'
ASSET_MANAGEMENT_DEPARTMENT_COLUMN = '부서'
ASSET_MANAGEMENT_ASSET_NO_COLUMN = '자산관리번호'
ASSET_MANAGEMENT_AMOUNT_COLUMN = '구입금액\n(VAT별도)'
ASSET_MANAGEMENT_STATUS_COLUMN = '물품상태\n(상/중/하)'
ASSET_MANAGEMENT_CATEGORY_COLUMN = '물품\n분류번호'
ASSET_MANAGEMENT_ITEM_NAME_COLUMN = '물품품목명'
ASSET_MANAGEMENT_STATE_TYPE_COLUMN = '상태 구분'
ASSET_MANAGEMENT_LOCATION_COLUMN = '설치장소'
ASSET_MANAGEMENT_USER_COLUMN = '사용자'
ASSET_MANAGEMENT_STATE_TYPE_OPTIONS = ['사용', '유후', '불용', '폐기대상']
ASSET_MANAGEMENT_FILTER_FIELDS = [
    ('department', ASSET_MANAGEMENT_DEPARTMENT_COLUMN),
    ('category', ASSET_MANAGEMENT_CATEGORY_COLUMN),
    ('item_name', ASSET_MANAGEMENT_ITEM_NAME_COLUMN),
    ('location', ASSET_MANAGEMENT_LOCATION_COLUMN),
    ('state_type', ASSET_MANAGEMENT_STATE_TYPE_COLUMN),
    ('status', ASSET_MANAGEMENT_STATUS_COLUMN),
    ('user', ASSET_MANAGEMENT_USER_COLUMN),
]
ASSET_MANAGEMENT_FIXED_HEADERS = [
    '부서',
    '일렬\n번호',
    '자산관리번호',
    '물품\n분류번호',
    '물품품목명',
    '품명/규격',
    '제조사\n구입처',
    '수량',
    '구입금액\n(VAT별도)',
    '구입날짜',
    '물품상태\n(상/중/하)',
    '상태 구분',
    '운용부서',
    '설치장소',
    '사용자',
    '정',
    '부',
    '비고',
    '재물조사 및 필증부착일',
    '1. 실물확인',
    '2. 파손유무',
    '3. 필증부착여부',
    '4. 특이사항',
    '폐기유무',
    '폐기사유',
    '백신설치 유무',
]
ASSET_MANAGEMENT_DEPARTMENT_OPTIONS = [
    '임원실',
    '기업부설연구소',
    '경영본부',
    '경영지원부',
    '총무부',
    '사업본부',
    'GIS사업부',
    '공간정보사업부',
    '영업본부',
    '공공사업부',
    '공정관리부',
    '비아이티',
    'BIT',
    'BIT 공정관리부',
]


def _normalize_asset_department(value: object) -> str:
    text = str(value or '').strip()
    return re.sub(r'^\d+\.', '', text).strip()


# Mapping from normalized department name -> short asset prefix used in `자산관리번호`
DEPARTMENT_ASSET_PREFIXES: dict[str, str] = {
    '경영지원부': '경영',
    '총무부': '총무',
    '공공사업부': '공공',
    '공정관리부': '공정',
    'GIS사업부': 'GIS',
    '공간정보사업부': '공간',
    '기업부설연구소': '연구소',
}


def _get_asset_header_label(header: str) -> str:
    label = str(header or '').replace('\n', ' ').strip()
    if header == ASSET_MANAGEMENT_STATUS_COLUMN:
        return '물품상태'
    
    # [추가된 부분] 내부 데이터는 '수량'이지만, 화면에 그릴 땐 '수량(개)'로 출력
    if header == '수량':
        return '수량(개)'
        
    return label


def _get_asset_header_labels(headers: list[str]) -> dict[str, str]:
    return {header: _get_asset_header_label(header) for header in headers}


def _normalize_asset_column_value(value: object) -> str:
    return str(value or '').strip()


def _normalize_asset_filter_value(filter_key: str, value: object) -> str:
    if filter_key == 'department':
        return _normalize_asset_department(value)
    return _normalize_asset_column_value(value)


def _normalize_asset_selected_filters(raw_filters: dict[str, object] | None) -> dict[str, str]:
    normalized_filters: dict[str, str] = {}
    raw_filters = raw_filters or {}
    for filter_key, _ in ASSET_MANAGEMENT_FILTER_FIELDS:
        normalized_filters[filter_key] = _normalize_asset_filter_value(filter_key, raw_filters.get(filter_key, ''))
    return normalized_filters


def _get_asset_selected_filters(args) -> dict[str, str]:
    return _normalize_asset_selected_filters(
        {filter_key: args.get(filter_key, '', type=str) for filter_key, _ in ASSET_MANAGEMENT_FILTER_FIELDS}
    )


def _apply_asset_management_filters(rows: list[dict[str, object]], selected_filters: dict[str, str]) -> list[dict[str, object]]:
    filtered_rows = rows
    for filter_key, column_name in ASSET_MANAGEMENT_FILTER_FIELDS:
        selected_value = selected_filters.get(filter_key, '')
        if not selected_value:
            continue
        filtered_rows = [
            row for row in filtered_rows
            if _normalize_asset_filter_value(filter_key, row.get(column_name, '')) == selected_value
        ]
    return filtered_rows


def _sort_asset_management_rows(rows: list[dict[str, object]], selected_filters: dict[str, str] | None = None) -> list[dict[str, object]]:
    selected_filters = selected_filters or {}
    indexed_rows = list(enumerate(rows))

    def sort_key(item: tuple[int, dict[str, object]]) -> tuple[int, int, str, str, int, str, int]:
        index, row = item
        department_text = str(row.get(ASSET_MANAGEMENT_DEPARTMENT_COLUMN, '') or '').strip()
        match = re.match(r'^(\d+)\.', department_text)
        department_order = int(match.group(1)) if match else 999999
        normalized_department = _normalize_asset_department(department_text)
        category = str(row.get(ASSET_MANAGEMENT_CATEGORY_COLUMN, '') or '').strip()
        asset_no = str(row.get(ASSET_MANAGEMENT_ASSET_NO_COLUMN, '') or '').strip()
        
        # [추가됨] 자산관리번호에 'BIT'가 포함되어 있으면 그룹 내 가장 마지막(1)으로 정렬
        is_bit = 1 if 'BIT' in asset_no else 0
        
        return (0 if match else 1, department_order, normalized_department, category, is_bit, asset_no, index)

    return [row for _, row in sorted(indexed_rows, key=sort_key)]


def _sort_asset_filter_options(values: set[str]) -> list[str]:
    return sorted(values, key=lambda item: (item.casefold(), item))


def _build_asset_filter_options(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    options_by_key: dict[str, set[str]] = {filter_key: set() for filter_key, _ in ASSET_MANAGEMENT_FILTER_FIELDS}

    for row in rows:
        for filter_key, column_name in ASSET_MANAGEMENT_FILTER_FIELDS:
            normalized_value = _normalize_asset_filter_value(filter_key, row.get(column_name, ''))
            if normalized_value:
                options_by_key[filter_key].add(normalized_value)

    department_values = []
    seen_departments: set[str] = set()
    for department_name in ASSET_MANAGEMENT_DEPARTMENT_OPTIONS:
        normalized_department = _normalize_asset_department(department_name)
        if normalized_department in options_by_key['department'] and normalized_department not in seen_departments:
            department_values.append(normalized_department)
            seen_departments.add(normalized_department)
    for department_name in _sort_asset_filter_options(options_by_key['department']):
        if department_name not in seen_departments:
            department_values.append(department_name)

    filter_options: list[dict[str, object]] = []
    for filter_key, column_name in ASSET_MANAGEMENT_FILTER_FIELDS:
        if filter_key == 'department':
            option_values = department_values
        elif filter_key == 'state_type':
            option_values = list(ASSET_MANAGEMENT_STATE_TYPE_OPTIONS)
        else:
            option_values = _sort_asset_filter_options(options_by_key[filter_key])
        filter_options.append(
            {
                'key': filter_key,
                'column': column_name,
                'label': _get_asset_header_label(column_name),
                'options': option_values,
            }
        )
    return filter_options


def _build_asset_filter_query_params(selected_filters: dict[str, str]) -> dict[str, str]:
    return {filter_key: value for filter_key, value in selected_filters.items() if value}


def _is_asset_modified_only_enabled(value: object) -> bool:
    return str(value or '').strip().lower() in {'1', 'true', 'y', 'yes', 'on'}


def _build_asset_export_workbook(year: int, headers: list[str], rows: list[dict[str, object]]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f'{year} 재물관리'

    header_labels = _get_asset_header_labels(headers)
    worksheet.append([header_labels.get(header, header) for header in headers])

    for row in rows:
        worksheet.append([_normalize_asset_column_value(row.get(header, '')) for header in headers])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _get_asset_row_identity(row_data: dict[str, object]) -> tuple[str, str]:
    return (
        _normalize_asset_department(row_data.get(ASSET_MANAGEMENT_DEPARTMENT_COLUMN, '')),
        _normalize_asset_column_value(row_data.get(ASSET_MANAGEMENT_ASSET_NO_COLUMN, '')),
    )


def _get_asset_history_title(row_data: dict[str, object], change_count: int | None = None) -> str:
    department = _normalize_asset_department(row_data.get(ASSET_MANAGEMENT_DEPARTMENT_COLUMN, '')) or '-'
    asset_no = _normalize_asset_column_value(row_data.get(ASSET_MANAGEMENT_ASSET_NO_COLUMN, '')) or '-'
    if change_count is None:
        return f'{department} - {asset_no}'
    return f'{department} - {asset_no} - {change_count}건'


def _parse_asset_no_suffix(asset_no: str) -> tuple[str, int, int] | None:
    text = str(asset_no or '').strip()
    match = re.search(r'^(.*?)(\d+)$', text)
    if not match:
        return None
    prefix = match.group(1)
    suffix = match.group(2)
    try:
        return prefix, int(suffix), len(suffix)
    except ValueError:
        return None


def _derive_asset_no_for_department_move(
    target_department: str,
    category: str,
    original_asset_no: str,
    rows: list[dict[str, object]],
) -> str:
    target_department = _normalize_asset_department(target_department)
    category = _normalize_asset_column_value(category)
    original_asset_no = _normalize_asset_column_value(original_asset_no)

    same_group_asset_nos = []
    for row in rows:
        asset_no = _normalize_asset_column_value(row.get(ASSET_MANAGEMENT_ASSET_NO_COLUMN, ''))
        if asset_no == original_asset_no:
            continue
        if _normalize_asset_department(row.get(ASSET_MANAGEMENT_DEPARTMENT_COLUMN, '')) != target_department:
            continue
        if _normalize_asset_column_value(row.get(ASSET_MANAGEMENT_CATEGORY_COLUMN, '')) != category:
            continue
        parsed = _parse_asset_no_suffix(asset_no)
        if parsed is None:
            continue
        prefix, seq, width = parsed
        same_group_asset_nos.append((prefix, seq, width))

    mapped_prefix = DEPARTMENT_ASSET_PREFIXES.get(target_department, None)
    if same_group_asset_nos:
        # Use mapped prefix when available, otherwise keep detected prefix
        _, max_seq, width = max(same_group_asset_nos, key=lambda item: item[1])
        seq = str(max_seq + 1).zfill(width)
        if mapped_prefix and category:
            return f"{mapped_prefix} {category}-{seq}"
        if mapped_prefix:
            return f"{mapped_prefix} {seq}"
        # fallback to using detected prefix formatting
        prefix, _, _ = max(same_group_asset_nos, key=lambda item: item[1])
        return f"{prefix}{seq}"

    original_parsed = _parse_asset_no_suffix(original_asset_no)
    if original_parsed is not None:
        _, seq_val, width = original_parsed
        seq = "1".zfill(width)
        if mapped_prefix and category:
            return f"{mapped_prefix} {category}-{seq}"
        if mapped_prefix:
            return f"{mapped_prefix} {seq}"
        return original_asset_no

    return original_asset_no


def _get_asset_row_highlight_key(row_data: dict[str, object]) -> str:
    return _get_asset_history_title(row_data)


def _extract_asset_history_key(change_summary: str) -> str:
    summary = str(change_summary or '').strip()
    return re.sub(r'\s*-\s*\d+건\s*$', '', summary).strip()


def _group_asset_management_history(history_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}
    ordered_keys: list[str] = []

    for history_row in history_rows:
        base_key = _extract_asset_history_key(history_row.get('change_summary', ''))
        if not base_key:
            continue
        if base_key not in grouped:
            grouped[base_key] = {
                'base_key': base_key,
                'change_summary': '',
                'latest_changed_at': history_row.get('changed_at', ''),
                'latest_changed_by': history_row.get('changed_by', '-'),
                'records': [],
            }
            ordered_keys.append(base_key)

        record_lines = []
        change_details = history_row.get('change_details') or []
        if change_details:
            for detail in change_details:
                record_lines.append(
                    f"{detail.get('label', '-')}: {detail.get('before', '-') or '-'} -> {detail.get('after', '-') or '-'}"
                )
        else:
            record_lines.append(history_row.get('change_summary', '-') or '-')

        history_action = history_row.get('action_type', 'edit') or 'edit'
        grouped[base_key]['records'].append(
            {
                'action_type': history_action,
                'action_label': {
                    'add': '추가',
                    'delete': '삭제',
                    'edit': '수정',
                    'move_in': '부서 이동',
                    'move_out': '부서 이동',
                }.get(history_action, '수정'),
                'change_lines': record_lines,
                'changed_at': history_row.get('changed_at', ''),
                'changed_by': history_row.get('changed_by', '-'),
            }
        )

    grouped_rows = []
    for base_key in ordered_keys:
        group = grouped[base_key]
        group['change_summary'] = f"{base_key} - {len(group['records'])}건"
        grouped_rows.append(group)
    return grouped_rows


def _build_asset_change_details(before_row: dict[str, object], after_row: dict[str, object], headers: list[str]) -> list[dict[str, str]]:
    details: list[dict[str, str]] = []
    for header in headers:
        before_value = _normalize_asset_column_value(before_row.get(header, ''))
        after_value = _normalize_asset_column_value(after_row.get(header, ''))
        
        # ▼ 변경된 비교 로직 ▼
        is_same = False
        
        # 1. 완벽히 일치하면 패스
        if before_value == after_value:
            is_same = True
            
        # 2. '상태 구분' 컬럼에서 빈칸과 '-'는 같은 것으로 취급
        elif header == ASSET_MANAGEMENT_STATE_TYPE_COLUMN:
            if before_value in ('', '-') and after_value in ('', '-'):
                is_same = True
                
        # 3. 금액 컬럼은 콤마(,)를 모두 제거한 상태에서 비교
        elif header in (ASSET_MANAGEMENT_AMOUNT_COLUMN, '수량'):
            if before_value.replace(',', '') == after_value.replace(',', ''):
                is_same = True

        if is_same:
            continue
        # ▲ 변경된 비교 로직 끝 ▲

        details.append(
            {
                'column': header,
                'label': _get_asset_header_label(header),
                'before': before_value,
                'after': after_value,
            }
        )
    return details


def _is_asset_blank_row(row_data: dict[str, object], headers: list[str]) -> bool:
    for header in headers:
        if _normalize_asset_column_value(row_data.get(header, '')):
            return False
    return True


def _build_asset_added_details(after_row: dict[str, object], headers: list[str]) -> list[dict[str, str]]:
    details: list[dict[str, str]] = []
    for header in headers:
        after_value = _normalize_asset_column_value(after_row.get(header, ''))
        if not after_value:
            continue
        details.append(
            {
                'column': header,
                'label': _get_asset_header_label(header),
                'before': '',
                'after': after_value,
            }
        )
    if not details:
        details.append(
            {
                'column': '__row_added__',
                'label': '행추가',
                'before': '',
                'after': '새 행 추가',
            }
        )
    return details


def _build_asset_deleted_details(before_row: dict[str, object], headers: list[str]) -> list[dict[str, str]]:
    details: list[dict[str, str]] = []
    for header in headers:
        before_value = _normalize_asset_column_value(before_row.get(header, ''))
        if not before_value:
            continue
        details.append(
            {
                'column': header,
                'label': _get_asset_header_label(header),
                'before': before_value,
                'after': '',
            }
        )
    if not details:
        details.append(
            {
                'column': '__row_deleted__',
                'label': '행삭제',
                'before': '기존 행 삭제',
                'after': '',
            }
        )
    return details


def _ensure_asset_management_tables(cursor) -> None:
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {ASSET_MANAGEMENT_IMPORT_TABLE} (
            id INT AUTO_INCREMENT PRIMARY KEY,
            asset_year INT NOT NULL,
            source_filename VARCHAR(255) NOT NULL,
            sheet_name VARCHAR(255) NOT NULL,
            imported_by VARCHAR(120) NULL,
            headers_json LONGTEXT NOT NULL,
            row_count INT NOT NULL DEFAULT 0,
            imported_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            KEY idx_asset_management_year (asset_year),
            KEY idx_asset_management_imported_at (imported_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {ASSET_MANAGEMENT_ROW_TABLE} (
            id INT AUTO_INCREMENT PRIMARY KEY,
            import_id INT NOT NULL,
            row_index INT NOT NULL,
            row_data LONGTEXT NOT NULL,
            created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            CONSTRAINT fk_asset_management_rows_import
                FOREIGN KEY (import_id) REFERENCES {ASSET_MANAGEMENT_IMPORT_TABLE}(id)
                ON DELETE CASCADE,
            UNIQUE KEY uq_asset_management_row (import_id, row_index)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {ASSET_MANAGEMENT_HISTORY_TABLE} (
            id INT AUTO_INCREMENT PRIMARY KEY,
            asset_year INT NOT NULL,
            action_type VARCHAR(40) NOT NULL,
            change_summary VARCHAR(255) NOT NULL,
            change_detail LONGTEXT NULL,
            changed_by VARCHAR(120) NULL,
            changed_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            KEY idx_asset_management_history_year (asset_year),
            KEY idx_asset_management_history_changed_at (changed_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    cursor.execute(f"SHOW COLUMNS FROM {ASSET_MANAGEMENT_HISTORY_TABLE} LIKE 'change_detail'")
    if not cursor.fetchone():
        cursor.execute(
            f"ALTER TABLE {ASSET_MANAGEMENT_HISTORY_TABLE} ADD COLUMN change_detail LONGTEXT NULL AFTER change_summary"
        )


def _record_asset_management_history(cursor, year: int, action_type: str, change_summary: str, change_detail=None) -> None:
    cursor.execute(
        f"""
        INSERT INTO {ASSET_MANAGEMENT_HISTORY_TABLE}
            (asset_year, action_type, change_summary, change_detail, changed_by)
        VALUES
            (%s, %s, %s, %s, %s)
        """,
        (
            year,
            action_type,
            change_summary,
            json.dumps(change_detail, ensure_ascii=False) if change_detail is not None else None,
            (session.get('user', {}) or {}).get('name', ''),
        ),
    )


def _serialize_asset_cell_value(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, Decimal):
        text = format(value, 'f')
        return text.rstrip('0').rstrip('.') if '.' in text else text
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value if isinstance(value, (int, float)) else str(value).strip()


def _build_asset_headers(raw_headers) -> list[str]:
    headers = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw_headers, start=1):
        header = str(value).strip() if value is not None and str(value).strip() else f'컬럼{index}'
        duplicate_count = seen.get(header, 0) + 1
        seen[header] = duplicate_count
        headers.append(header if duplicate_count == 1 else f'{header}_{duplicate_count}')
    return headers


def _extract_asset_management_rows_from_excel(file_stream):
    workbook = load_workbook(file_stream, data_only=True)
    worksheet = workbook.active

    header_row_index = 3
    if worksheet.max_row < header_row_index:
        raise ValueError('엑셀에서 헤더 행을 찾지 못했습니다.')

    header_values = [cell.value for cell in worksheet[header_row_index]]
    if not any(cell not in (None, '') for cell in header_values):
        raise ValueError('엑셀 3행에 컬럼명이 없습니다.')

    headers = list(ASSET_MANAGEMENT_FIXED_HEADERS)
    rows = []
    
    # 💡 values_only=False 로 변경하여 셀의 색상 속성을 읽어옵니다.
    for row_index, row_cells in enumerate(
        worksheet.iter_rows(min_row=header_row_index + 1, max_row=worksheet.max_row, values_only=False),
        start=1,
    ):
        # cell.value 로 값 추출
        serialized = [_serialize_asset_cell_value(cell.value) for cell in row_cells[:len(headers)]]
        if not any(str(cell).strip() for cell in serialized if cell not in (None, '')):
            continue

        row_data = {}
        for column_index, header in enumerate(headers):
            row_data[header] = serialized[column_index] if column_index < len(serialized) else ''
            
        # 💡 [추가된 로직] 행의 배경색(노란색, 빨간색) 읽어오기
        row_color_class = ''
        for cell in row_cells[:len(headers)]:
            fill = cell.fill
            if fill and fill.patternType == 'solid' and fill.start_color and fill.start_color.type == 'rgb':
                rgb = str(fill.start_color.rgb).upper()
                
                # 노란색 계열 확인 -> 부서 이동 들어옴(moved-in, 노랑) 처리
                if rgb.endswith('FFFF00') or rgb in ('FFFFFF99', 'FFFFFFCC', 'FFFFE699', 'FFFFFF66'):
                    row_color_class = 'asset-row-moved-in'
                    break
                # 빨간색 계열 확인 -> 부서 이동 나감(moved-out, 빨강) 처리
                elif rgb.endswith('FF0000') or rgb in ('FFFF9999', 'FFFFC7CE', 'FFFFCCCC', 'FFFFE5E5'):
                    row_color_class = 'asset-row-moved-out'
                    break
        
        # 색상이 존재하면 row_data에 특별한 키값으로 저장
        if row_color_class:
            row_data['__row_color_class'] = row_color_class

        rows.append({'row_index': row_index, 'row_data': row_data})

    if not rows:
        raise ValueError('엑셀에 저장할 데이터 행이 없습니다.')

    return worksheet.title or 'Sheet1', headers, rows


def _load_asset_management_view_data(cursor, year: int):
    _ensure_asset_management_tables(cursor)
    cursor.execute(
        f"""
        SELECT id, asset_year, source_filename, sheet_name, imported_by, headers_json, row_count, imported_at
        FROM {ASSET_MANAGEMENT_IMPORT_TABLE}
        WHERE asset_year = %s
        ORDER BY imported_at ASC, id ASC
        """,
        (year,),
    )
    import_rows = cursor.fetchall() or []
    if not import_rows:
        return None, [], []

    import_ids = []
    for import_row in import_rows:
        import_ids.append(import_row['id'])
    headers = list(ASSET_MANAGEMENT_FIXED_HEADERS)

    placeholders = ','.join(['%s'] * len(import_ids))

    cursor.execute(
        f"""
        SELECT import_id, row_index, row_data
        FROM {ASSET_MANAGEMENT_ROW_TABLE}
        WHERE import_id IN ({placeholders})
        ORDER BY import_id ASC, row_index ASC, id ASC
        """,
        tuple(import_ids),
    )
    rows = []
    for original_row_index, row in enumerate(cursor.fetchall() or []):
        try:
            row_data = json.loads(row.get('row_data') or '{}')
        except json.JSONDecodeError:
            row_data = {}
        if not isinstance(row_data, dict):
            row_data = {}
        normalized_row = {header: row_data.get(header, '') for header in headers}
        
        #색상유지
        if '__row_color_class' in row_data:
            normalized_row['__row_color_class'] = row_data['__row_color_class']
        # Treat stored '-' for 상태 구분 as blank for display
        try:
            if ASSET_MANAGEMENT_STATE_TYPE_COLUMN in normalized_row and str(normalized_row.get(ASSET_MANAGEMENT_STATE_TYPE_COLUMN) or '').strip() == '-':
                normalized_row[ASSET_MANAGEMENT_STATE_TYPE_COLUMN] = ''
        except Exception:
            pass
        normalized_row['__original_row_index'] = original_row_index
        rows.append(normalized_row)

    latest_import = import_rows[-1]
    imported_at = latest_import.get('imported_at')
    imported_at_label = imported_at.strftime('%Y-%m-%d %H:%M:%S') if isinstance(imported_at, datetime) else ''
    meta = {
        'year': latest_import.get('asset_year'),
        'source_filename': latest_import.get('source_filename') or '',
        'sheet_name': latest_import.get('sheet_name') or '',
        'imported_by': latest_import.get('imported_by') or '-',
        'imported_at': imported_at_label,
        'row_count': len(rows),
    }
    return meta, headers, rows


def _load_asset_management_history(cursor, year: int, action_types: list[str] | None = None):
    _ensure_asset_management_tables(cursor)
    query = f"""
        SELECT id, action_type, change_summary, change_detail, changed_by, changed_at
        FROM {ASSET_MANAGEMENT_HISTORY_TABLE}
        WHERE asset_year = %s
    """
    params: list[object] = [year]
    if action_types:
        placeholders = ','.join(['%s'] * len(action_types))
        query += f" AND action_type IN ({placeholders})"
        params.extend(action_types)
    query += ' ORDER BY changed_at DESC, id DESC'
    cursor.execute(query, tuple(params))
    history_rows = []
    for row in cursor.fetchall() or []:
        changed_at = row.get('changed_at')
        try:
            detail_items = json.loads(row.get('change_detail') or '[]')
        except json.JSONDecodeError:
            detail_items = []
        if not isinstance(detail_items, list):
            detail_items = []
        history_rows.append(
            {
                'id': row.get('id'),
                'action_type': row.get('action_type') or '',
                'change_summary': row.get('change_summary') or '',
                'change_details': detail_items,
                'changed_by': row.get('changed_by') or '-',
                'changed_at': changed_at.strftime('%Y-%m-%d %H:%M:%S') if isinstance(changed_at, datetime) else '',
            }
        )
    return history_rows


def _has_admin_access() -> bool:
    user = session.get('user', {}) or {}
    if (user.get('auth') or '').strip() == '관리자':
        return True
    try:
        return int(user.get('adminAUTH', user.get('adminauth', 0)) or 0) == 1
    except (TypeError, ValueError):
        return False


def _format_project_raw_value(column_name: str, value):
    if value is None:
        return ''

    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')

    if column_name in PROJECT_RAW_BOOLEAN_COLUMNS:
        return 'Y' if int(value or 0) == 1 else 'N'

    if isinstance(value, Decimal):
        text = format(value, 'f')
        return text.rstrip('0').rstrip('.') if '.' in text else text

    if isinstance(value, int):
        if column_name in PROJECT_RAW_MONEY_COLUMNS:
            return f'{value:,}'
        return str(value)

    if isinstance(value, float):
        if column_name in PROJECT_RAW_MONEY_COLUMNS:
            return f'{value:,.0f}'
        return format(value, '.6g')

    return str(value)


def _build_project_raw_column_names(project_column_names: list[str]) -> list[str]:
    column_names = []
    reference_inserted = False
    for column_name in project_column_names:
        if column_name in PROJECT_RAW_EXCLUDED_COLUMNS:
            continue
        if column_name in PROJECT_RAW_REFERENCE_COLUMNS:
            if not reference_inserted:
                column_names.append('referenceProjects')
                reference_inserted = True
            continue
        column_names.append(column_name)
    return column_names


def _fetch_contract_linked_tables(cursor) -> list[tuple[str, str]]:
    cursor.execute(
        """
        SELECT table_name, column_name
        FROM information_schema.columns
        WHERE table_schema = DATABASE()
          AND LOWER(REPLACE(column_name, ' ', '')) IN ('contractcode', 'contract_code')
        ORDER BY table_name, ordinal_position
        """
    )
    linked_tables = []
    seen_tables = set()
    for row in cursor.fetchall() or []:
        table_name = row.get('table_name', row.get('TABLE_NAME'))
        column_name = row.get('column_name', row.get('COLUMN_NAME'))
        if not table_name or not column_name or table_name == 'projects' or table_name in seen_tables:
            continue
        seen_tables.add(table_name)
        linked_tables.append((table_name, column_name))
    return linked_tables


def _fetch_related_table_columns(cursor, contract_codes: list[str]) -> list[dict]:
    if not contract_codes:
        return []

    placeholders = ','.join(['%s'] * len(contract_codes))
    related_columns = []
    for table_name, contract_column in _fetch_contract_linked_tables(cursor):
        cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
        table_columns = [row.get('Field') for row in (cursor.fetchall() or []) if row.get('Field')]
        if not table_columns or contract_column not in table_columns:
            continue

        select_columns_sql = ', '.join(f'`{column_name}`' for column_name in table_columns)
        cursor.execute(
            f"""
            SELECT {select_columns_sql}
            FROM `{table_name}`
            WHERE `{contract_column}` IN ({placeholders})
            ORDER BY `{contract_column}`
            """,
            tuple(contract_codes),
        )
        grouped_rows = {}
        for row in cursor.fetchall() or []:
            contract_code = row.get(contract_column)
            if not contract_code:
                continue
            grouped_rows.setdefault(contract_code, []).append(row)

        table_label = PROJECT_RAW_RELATED_TABLE_LABELS.get(table_name, table_name)
        for column_name in table_columns:
            if column_name == contract_column:
                continue

            related_columns.append(
                {
                    'name': f'related__{table_name}__{column_name}',
                    'label': f'{table_label}.{column_name}',
                    'values': {
                        contract_code: ' || '.join(
                            f'[{index + 1}] {formatted_value}'
                            for index, formatted_value in enumerate(
                                [
                                    _format_project_raw_value(column_name, item.get(column_name))
                                    for item in items
                                    if _format_project_raw_value(column_name, item.get(column_name)) != ''
                                ]
                            )
                        )
                        for contract_code, items in grouped_rows.items()
                    },
                }
            )
    return related_columns


@bp.route('/api/get_projects/')
def get_projects():
    """선택된 연도의 프로젝트 데이터를 JSON으로 반환 (페이지네이션 적용)"""
    db = create_connection()
    if db is None:
        return jsonify({"error": "Database connection could not be established"}), 500

    cursor = None
    try:
        cursor = db.cursor(dictionary=True)
        year = request.args.get('year', type=int)
        page = request.args.get('page', 1, type=int)
        per_page = 20

        if not year:
            return jsonify({"error": "Year is required"}), 400

        cursor.execute(
            """
                SELECT COUNT(*) AS count FROM Projects
            WHERE YEAR(StartDate) = %s AND ContractCode NOT LIKE '%%검토%%'
            """,
            (year,),
        )
        total_projects = cursor.fetchone()['count']
        total_pages = max(1, (total_projects + per_page - 1) // per_page)

        if page > total_pages or page < 1:
            page = 1

        offset = (page - 1) * per_page

        cursor.execute(
            """
                SELECT ProjectID, ProjectName, ContractCode, orderPlace, yearProject, outsourcingCheck, project_status
                FROM Projects
                WHERE YEAR(StartDate) = %s
            AND ContractCode NOT LIKE '%%검토%%'
                ORDER BY ContractCode DESC
            LIMIT %s OFFSET %s
            """,
            (year, per_page, offset),
        )
        projects = cursor.fetchall()

        try:
            contract_codes = [
                p.get('ContractCode') for p in projects if isinstance(p, dict) and p.get('ContractCode')
            ]
            progress_map = calc_progress_bulk(contract_codes)
            for proj in projects:
                if isinstance(proj, dict):
                    code = proj.get('ContractCode')
                    proj['progress'] = progress_map.get(code, 0.0)

            risk_map = {}
            if contract_codes:
                try:
                    has_division = False
                    try:
                        cursor.execute("SHOW COLUMNS FROM project_risks LIKE 'division'")
                        has_division = cursor.fetchone() is not None
                    except Exception:
                        has_division = False

                    fmt = ','.join(['%s'] * len(contract_codes))
                    if has_division:
                        cursor.execute(
                            f"""
                            SELECT DISTINCT contractcode FROM project_risks
                            WHERE contractcode IN ({fmt})
                              AND (division IS NULL OR division <> '완료')
                            """,
                            tuple(contract_codes),
                        )
                    else:
                        cursor.execute(
                            f"""
                            SELECT DISTINCT contractcode FROM project_risks
                            WHERE contractcode IN ({fmt})
                            """,
                            tuple(contract_codes),
                        )
                    for r in cursor.fetchall() or []:
                        code = r.get('contractcode') if isinstance(r, dict) else None
                        if code:
                            risk_map[code] = True
                except Exception:
                    risk_map = {}
            for proj in projects:
                if isinstance(proj, dict):
                    code = proj.get('ContractCode')
                    proj['has_risk'] = bool(risk_map.get(code))
        except Exception:
            pass

        return jsonify({
            'projects': projects,
            'total_pages': total_pages,
            'current_page': page,
        })

    except mysql.connector.Error as e:
        print(f"Error executing SQL query: {e}")
        return jsonify({"error": "Failed to fetch data"}), 500
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


@bp.route('/PMS_Business/<int:year>')
def business(year: int):
    db = create_connection()
    if db is None:
        return jsonify({"error": "Database connection could not be established"}), 500

    cursor = None
    try:
        cursor = db.cursor(dictionary=True)

        cursor.execute("SELECT DISTINCT YEAR(StartDate) as year FROM Projects ORDER BY year DESC")
        available_years = [row['year'] for row in cursor.fetchall()]

        cursor.execute(
            """
            SELECT ProjectID, ProjectName, ContractCode, orderPlace, yearProject, outsourcingCheck 
            FROM Projects 
            WHERE YEAR(StartDate) = %s 
            AND ContractCode NOT LIKE '%%검토%%'
            ORDER BY ContractCode DESC
            """,
            (year,),
        )
        projects = cursor.fetchall()

        cursor.execute(
            """
            SELECT ProjectID, ProjectName, ContractCode, orderPlace, yearProject, outsourcingCheck 
            FROM Projects 
            WHERE ContractCode NOT LIKE '%%검토%%'
            ORDER BY ContractCode DESC
            """
        )
        all_projects = cursor.fetchall()

        cursor.execute(
            """
             SELECT userID, Name, Department, Position, EmpNo, JoinDate, Phone, Auth, note,
                 COALESCE(adminAUTH, 0) AS adminAUTH,
                   COALESCE(dataauth, 0)   AS dataauth,
                   COALESCE(reportAUTH, 0) AS reportAUTH,
                 COALESCE(meetingAuth, 0) AS meetingAuth,
                   COALESCE(projectAUTH, 0) AS projectAUTH
            FROM users
            """
        )
        users = cursor.fetchall()

        for user in users:
            if user.get('note') is None:
                user['note'] = ''

    except mysql.connector.Error as e:
        print(f"Error executing SQL query: {e}")
        return jsonify({"error": "Failed to fetch data"}), 500
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

    return render_template(
        'PMS_Business_Year.html',
        projects=projects,
        ALL_PROJECT=all_projects,
        selected_year=year,
        available_years=available_years,
        users=users,
        Auth=session.get('user', {}).get('auth', '비회원'),
    )



def _calculate_asset_summary_and_aggregate(rows):
    # Departments in correct order
    DEPARTMENTS = [
        '1.경영지원부',
        '2.총무부',
        '3.공공사업부',
        '4.공정관리부',
        '5.GIS사업부',
        '6.공간정보사업부',
        '7.기업부설연구소'
    ]
    
    # Normalizers
    def norm_dept(d):
        d_str = str(d or '').strip()
        if '경영지원' in d_str: return '1.경영지원부'
        if '총무' in d_str: return '2.총무부'
        if '공공' in d_str: return '3.공공사업부'
        if '공정' in d_str: return '4.공정관리부'
        if 'GIS' in d_str: return '5.GIS사업부'
        if '공간정보' in d_str: return '6.공간정보사업부'
        if '연구소' in d_str or '기업부설연구소' in d_str: return '7.기업부설연구소'
        return d_str

    def norm_status(s):
        s_str = str(s or '').strip()
        if '사용' in s_str: return '1.사용'
        if '유휴' in s_str or '유후' in s_str: return '2.유휴'
        if '불용' in s_str: return '3.불용'
        if '폐기' in s_str: return '4.폐기대상'
        if '분실' in s_str: return '5.분실'
        return s_str

    # 1. Filter active rows (실물확인 == '○' / 'o' / 'O')
    active_rows = []
    for r in rows:
        chk = str(r.get('1. 실물확인') or '').strip().upper()
        if chk in ('○', 'O', 'O'):
            active_rows.append(r)

    # 2. Calculate Aggregate Details (집계내역)
    aggregate_grouped = {}
    for r in active_rows:
        cat_no = str(r.get('물품\n분류번호') or '').strip()
        item_name = str(r.get('물품품목명') or '').strip()
        if not cat_no and not item_name:
            continue
            
        qty_str = str(r.get('수량') or '0').strip()
        try:
            qty = float(qty_str) if '.' in qty_str else int(qty_str)
        except ValueError:
            qty = 0
            
        dept = norm_dept(r.get('부서'))
        if dept not in DEPARTMENTS:
            continue
            
        key = (cat_no, item_name)
        if key not in aggregate_grouped:
            aggregate_grouped[key] = {d: 0 for d in DEPARTMENTS}
        aggregate_grouped[key][dept] += qty

    sorted_keys = sorted(aggregate_grouped.keys(), key=lambda x: (x[0], x[1]))
    aggregate_rows = []
    for key in sorted_keys:
        cat_no, item_name = key
        row_map = aggregate_grouped[key]
        row_sum = sum(row_map.values())
        aggregate_rows.append({
            'category_no': cat_no,
            'item_name': item_name,
            'values': {d: row_map[d] for d in DEPARTMENTS},
            'total': row_sum
        })

    # 3. Calculate Summary Table (총괄표)
    status_keys = ['1.사용', '2.유휴', '3.불용', '4.폐기대상', '5.분실']
    status_totals = {st: {d: 0 for d in DEPARTMENTS} for st in status_keys}
    status_item_breakdown = {st: {} for st in status_keys}

    for r in active_rows:
        status = norm_status(r.get('상태 구분'))
        if status not in status_totals:
            continue
            
        dept = norm_dept(r.get('부서'))
        if dept not in DEPARTMENTS:
            continue
            
        qty_str = str(r.get('수량') or '0').strip()
        try:
            qty = float(qty_str) if '.' in qty_str else int(qty_str)
        except ValueError:
            qty = 0
            
        item_name = str(r.get('물품품목명') or '').strip()
        
        status_totals[status][dept] += qty
        
        if item_name:
            item_map = status_item_breakdown[status]
            if item_name not in item_map:
                item_map[item_name] = {d: 0 for d in DEPARTMENTS}
            item_map[item_name][dept] += qty

    # Format Summary Table Rows
    summary_data = []
    for st in status_keys:
        st_label = {
            '1.사용': '사용',
            '2.유휴': '유휴',
            '3.불용': '불용(파손)',
            '4.폐기대상': '폐기대상',
            '5.분실': '분실'
        }.get(st, st)
        
        st_map = status_totals[st]
        st_sum = sum(st_map.values())
        
        summary_data.append({
            'type': 'header',
            'label': st_label,
            'values': {d: st_map[d] for d in DEPARTMENTS},
            'total': st_sum
        })
        
        # Breakdown rows for non-사용 statuses
        if st != '1.사용':
            item_map = status_item_breakdown[st]
            sorted_items = sorted(item_map.keys(), key=lambda x: x.casefold())
            for item in sorted_items:
                item_vals = item_map[item]
                item_sum = sum(item_vals.values())
                if item_sum > 0:
                    summary_data.append({
                        'type': 'item',
                        'label': item,
                        'values': {d: item_vals[d] for d in DEPARTMENTS},
                        'total': item_sum
                    })
                    
    return summary_data, aggregate_rows

@bp.route('/PMS_Business/asset-management/<int:year>')
def asset_management(year: int):
    if not _has_admin_access():
        return '권한이 없습니다.', 403

    db = create_connection()
    if db is None:
        return jsonify({'error': 'Database connection could not be established'}), 500

    cursor = None
    try:
        cursor = db.cursor(dictionary=True)
        asset_meta, columns, rows = _load_asset_management_view_data(cursor, year)
        
        # ▼ 추가된 부분: 총괄표 및 집계내역 계산 함수 호출
        summary_rows, aggregate_rows = _calculate_asset_summary_and_aggregate(rows)

        selected_filters = _get_asset_selected_filters(request.args)
        filter_options = _build_asset_filter_options(rows)
        rows = _apply_asset_management_filters(rows, selected_filters)
        history_rows = _load_asset_management_history(cursor, year)
        grouped_history_rows = _group_asset_management_history(history_rows)
        highlighted_row_keys = {}
        for item in history_rows:
            if not item.get('change_summary'):
                continue
            base_key = _extract_asset_history_key(item.get('change_summary', ''))
            action_type = item.get('action_type', 'edit') or 'edit'
            if base_key not in highlighted_row_keys or highlighted_row_keys[base_key] == 'asset-row-modified':
                highlighted_row_keys[base_key] = {
                    'move_out': 'asset-row-moved-out',
                    'move_in': 'asset-row-moved-in',
                    'add': 'asset-row-added',
                    'edit': 'asset-row-modified',
                    'delete': 'asset-row-deleted',
                }.get(action_type, 'asset-row-modified')
        show_modified_only = _is_asset_modified_only_enabled(request.args.get('show_modified', '', type=str))
        if show_modified_only:
            rows = [row for row in rows if _get_asset_row_highlight_key(row) in highlighted_row_keys]
        rows = _sort_asset_management_rows(rows, selected_filters)
        latest_modified_at = history_rows[0]['changed_at'] if history_rows else '-'
        filter_query_params = _build_asset_filter_query_params(selected_filters)
        if show_modified_only:
            filter_query_params['show_modified'] = '1'
            
        return render_template(
            'PMS_asset_management.html',
            year=year,
            columns=columns,
            rows=rows,
            total_count=len(rows),
            asset_meta=asset_meta,
            
            # ▼ 추가된 부분: HTML로 계산된 데이터 전달
            asset_summary_rows=summary_rows,
            asset_aggregate_rows=aggregate_rows,
            
            asset_history=grouped_history_rows,
            asset_history_count=len(grouped_history_rows),
            asset_latest_modified_at=latest_modified_at,
            asset_header_labels=_get_asset_header_labels(columns),
            asset_highlighted_row_keys=highlighted_row_keys,
            asset_filter_options=filter_options,
            asset_department_options=ASSET_MANAGEMENT_DEPARTMENT_OPTIONS,
            asset_department_column=ASSET_MANAGEMENT_DEPARTMENT_COLUMN,
            asset_asset_no_column=ASSET_MANAGEMENT_ASSET_NO_COLUMN,
            asset_amount_column=ASSET_MANAGEMENT_AMOUNT_COLUMN,
            asset_status_column=ASSET_MANAGEMENT_STATUS_COLUMN,
            selected_filters=selected_filters,
            selected_department=selected_filters.get('department', ''),
            asset_show_modified_only=show_modified_only,
            asset_page_url=url_for('business_year.asset_management', year=year, **filter_query_params),
            asset_page_base_url=url_for('business_year.asset_management', year=year),
            asset_frame_url=url_for('business_year.asset_management_frame', year=year, **filter_query_params),
            asset_export_url=url_for('business_year.export_asset_management_excel'),
            error=None,
        )
    except mysql.connector.Error as e:
        print(f'Error executing SQL query: {e}')
        return jsonify({'error': 'Failed to fetch asset management data'}), 500
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


@bp.route('/PMS_Business/asset-management/frame/<int:year>')
def asset_management_frame(year: int):
    if not _has_admin_access():
        return '권한이 없습니다.', 403

    db = create_connection()
    if db is None:
        return 'Database connection could not be established', 500

    cursor = None
    try:
        cursor = db.cursor(dictionary=True)
        _, columns, rows = _load_asset_management_view_data(cursor, year)
        selected_filters = _get_asset_selected_filters(request.args)
        rows = _apply_asset_management_filters(rows, selected_filters)
        history_rows = _load_asset_management_history(cursor, year)
        highlighted_row_keys = {}
        for item in history_rows:
            if not item.get('change_summary'):
                continue
            base_key = _extract_asset_history_key(item.get('change_summary', ''))
            action_type = item.get('action_type', 'edit') or 'edit'
            if base_key not in highlighted_row_keys or highlighted_row_keys[base_key] == 'asset-row-modified':
                highlighted_row_keys[base_key] = {
                    'move_out': 'asset-row-moved-out',
                    'move_in': 'asset-row-moved-in',
                    'add': 'asset-row-added',
                    'edit': 'asset-row-modified',
                    'delete': 'asset-row-deleted',
                }.get(action_type, 'asset-row-modified')
        if _is_asset_modified_only_enabled(request.args.get('show_modified', '', type=str)):
            rows = [row for row in rows if _get_asset_row_highlight_key(row) in highlighted_row_keys]
        rows = _sort_asset_management_rows(rows, selected_filters)
        return render_template(
            'PMS_asset_management_frame.html',
            year=year,
            columns=columns,
            rows=rows,
            asset_header_labels=_get_asset_header_labels(columns),
            asset_highlighted_row_keys=highlighted_row_keys,
            asset_department_column=ASSET_MANAGEMENT_DEPARTMENT_COLUMN,
            asset_asset_no_column=ASSET_MANAGEMENT_ASSET_NO_COLUMN,
            asset_amount_column=ASSET_MANAGEMENT_AMOUNT_COLUMN,
        )
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


@bp.route('/api/asset-management/upload', methods=['POST'])
def upload_asset_management_excel():
    if not _has_admin_access():
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    year = request.form.get('year', type=int)
    if not year:
        return jsonify({'success': False, 'message': '연도 정보가 없습니다.'}), 400

    file = request.files.get('file')
    if file is None or not file.filename:
        return jsonify({'success': False, 'message': '업로드할 엑셀 파일을 선택해 주세요.'}), 400
    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'message': '.xlsx 파일만 업로드 가능합니다.'}), 400

    try:
        sheet_name, headers, rows = _extract_asset_management_rows_from_excel(BytesIO(file.read()))
    except (InvalidFileException, BadZipFile, ValueError) as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        print(f'[ERROR] 재물관리 엑셀 처리 실패: {type(e).__name__}: {e}')
        return jsonify({'success': False, 'message': '엑셀 처리 중 오류가 발생했습니다.'}), 500

    db = create_connection()
    if db is None:
        return jsonify({'success': False, 'message': 'Database connection could not be established'}), 500

    cursor = None
    try:
        cursor = db.cursor(dictionary=True)
        _ensure_asset_management_tables(cursor)
        cursor.execute(
            f"""
            INSERT INTO {ASSET_MANAGEMENT_IMPORT_TABLE}
                (asset_year, source_filename, sheet_name, imported_by, headers_json, row_count)
            VALUES
                (%s, %s, %s, %s, %s, %s)
            """,
            (
                year,
                file.filename,
                sheet_name,
                (session.get('user', {}) or {}).get('name', ''),
                json.dumps(ASSET_MANAGEMENT_FIXED_HEADERS, ensure_ascii=False),
                len(rows),
            ),
        )
        import_id = cursor.lastrowid
        cursor.executemany(
            f"""
            INSERT INTO {ASSET_MANAGEMENT_ROW_TABLE}
                (import_id, row_index, row_data)
            VALUES
                (%s, %s, %s)
            """,
            [
                (import_id, row['row_index'], json.dumps(row['row_data'], ensure_ascii=False))
                for row in rows
            ],
        )
        db.commit()
        return jsonify(
            {
                'success': True,
                'message': f'{len(rows)}건을 재물관리 데이터로 누적 반영했습니다.',
                'row_count': len(rows),
                'sheet_name': sheet_name,
            }
        )
    except mysql.connector.Error as e:
        if db:
            db.rollback()
        print(f'[ERROR] 재물관리 업로드 저장 실패: {e}')
        return jsonify({'success': False, 'message': 'DB 저장 중 오류가 발생했습니다.'}), 500
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


@bp.route('/api/asset-management/save', methods=['POST'])
def save_asset_management_rows():
    if not _has_admin_access():
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    payload = request.get_json(silent=True) or {}
    year = payload.get('year')
    rows = payload.get('rows')
    selected_department = _normalize_asset_department(payload.get('department', ''))

    try:
        year = int(year)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': '연도 정보가 올바르지 않습니다.'}), 400

    if not isinstance(rows, list) or not rows:
        return jsonify({'success': False, 'message': '저장할 행 데이터가 없습니다.'}), 400

    normalized_rows = []
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        normalized_rows.append({'row_index': index, 'row_data': row})

    if not normalized_rows:
        return jsonify({'success': False, 'message': '저장 가능한 행 데이터가 없습니다.'}), 400

    db = create_connection()
    if db is None:
        return jsonify({'success': False, 'message': 'Database connection could not be established'}), 500

    cursor = None
    try:
        cursor = db.cursor(dictionary=True)
        _ensure_asset_management_tables(cursor)
        _, _, existing_rows = _load_asset_management_view_data(cursor, year)
        cursor.execute(
            f"""
            SELECT id, headers_json
            FROM {ASSET_MANAGEMENT_IMPORT_TABLE}
            WHERE asset_year = %s
            ORDER BY imported_at ASC, id ASC
            """,
            (year,),
        )
        import_rows = cursor.fetchall() or []
        if not import_rows:
            return jsonify({'success': False, 'message': '먼저 엑셀 데이터를 업로드해 주세요.'}), 400

        import_ids = []
        for import_row in import_rows:
            import_ids.append(import_row['id'])
        headers = list(ASSET_MANAGEMENT_FIXED_HEADERS)

        normalized_payload = []
        for row in normalized_rows:
            row_data = row['row_data']
            if headers:
                cleaned = {header: str(row_data.get(header, '') or '').strip() for header in headers}
            else:
                cleaned = {str(key): str(value or '').strip() for key, value in row_data.items()}
            original_row_index = row_data.get('__original_row_index')
            try:
                cleaned['__original_row_index'] = int(original_row_index)
            except (TypeError, ValueError):
                pass
            normalized_payload.append(cleaned)

        # Remove completely blank rows from the payload to avoid inserting garbage/empty rows
        normalized_payload = [r for r in normalized_payload if not _is_asset_blank_row(r, headers)]

        if selected_department:
            untouched_rows = [
                row for row in existing_rows
                if _normalize_asset_department(row.get(ASSET_MANAGEMENT_DEPARTMENT_COLUMN, '')) != selected_department
            ]
            before_rows = [
                row for row in existing_rows
                if _normalize_asset_department(row.get(ASSET_MANAGEMENT_DEPARTMENT_COLUMN, '')) == selected_department
            ]
            # Merge: preserve existing rows for the selected department unless the payload provides an explicit update.
            # Match by identity (department, asset_no) or by original index when provided.
            payload_by_identity = {}
            payload_by_index = {}
            for r in normalized_payload:
                identity = _get_asset_row_identity(r)
                if any(identity):
                    payload_by_identity[identity] = r
                orig_idx = r.get('__original_row_index')
                if isinstance(orig_idx, int):
                    payload_by_index[orig_idx] = r

            updated_before_rows = []
            matched_payload_identities = set()
            matched_payload_indexes = set()
            for b in before_rows:
                identity = _get_asset_row_identity(b)
                orig_idx = b.get('__original_row_index')
                new_row = None
                if any(identity) and identity in payload_by_identity:
                    new_row = payload_by_identity[identity]
                    matched_payload_identities.add(identity)
                elif isinstance(orig_idx, int) and orig_idx in payload_by_index:
                    new_row = payload_by_index[orig_idx]
                    matched_payload_indexes.add(orig_idx)

                if new_row is not None:
                    updated_before_rows.append(new_row)
                else:
                    updated_before_rows.append(b)

            # Append any new rows in payload that did not match existing ones (these are additions)
            new_rows = []
            for r in normalized_payload:
                identity = _get_asset_row_identity(r)
                orig_idx = r.get('__original_row_index')
                if (any(identity) and identity in matched_payload_identities) or (isinstance(orig_idx, int) and orig_idx in matched_payload_indexes):
                    continue
                # not matched -> treat as new addition
                new_rows.append(r)

            merged_payload = untouched_rows + updated_before_rows + new_rows
            # Ensure we do not insert fully blank rows into the DB
            merged_payload = [r for r in merged_payload if not _is_asset_blank_row(r, headers)]
        else:
            before_rows = existing_rows
            merged_payload = normalized_payload

        before_map = {
            _get_asset_row_identity(row): row
            for row in before_rows
            if any(_get_asset_row_identity(row))
        }
        before_index_map = {
            int(row.get('__original_row_index')): row
            for row in before_rows
            if isinstance(row.get('__original_row_index'), int)
        }
        matched_before_keys = set()
        matched_before_indexes: set[int] = set()
        change_entries = []
        moved_out_rows: list[dict[str, object]] = []
        for row in normalized_payload:
            identity = _get_asset_row_identity(row)
            if not any(identity):
                continue
            previous_row = before_map.get(identity)
            original_row_index = row.get('__original_row_index')
            if previous_row is None and isinstance(original_row_index, int):
                previous_row = before_index_map.get(original_row_index)
                if previous_row is not None:
                    if _is_asset_blank_row(previous_row, headers):
                        previous_row = None
                    else:
                        matched_before_indexes.add(original_row_index)
                        previous_identity = _get_asset_row_identity(previous_row)
                        if any(previous_identity):
                            matched_before_keys.add(previous_identity)
            
            if previous_row is not None:
                before_department = _normalize_asset_department(previous_row.get(ASSET_MANAGEMENT_DEPARTMENT_COLUMN, ''))
                after_department = _normalize_asset_department(row.get(ASSET_MANAGEMENT_DEPARTMENT_COLUMN, ''))
                if before_department and after_department and before_department != after_department:
                    moved_asset_no = _derive_asset_no_for_department_move(
                        after_department,
                        _normalize_asset_column_value(row.get(ASSET_MANAGEMENT_CATEGORY_COLUMN, '')),
                        _normalize_asset_column_value(row.get(ASSET_MANAGEMENT_ASSET_NO_COLUMN, '')),
                        existing_rows + normalized_payload,
                    )
                    row[ASSET_MANAGEMENT_ASSET_NO_COLUMN] = moved_asset_no

                    move_details = [
                        {
                            'column': ASSET_MANAGEMENT_DEPARTMENT_COLUMN,
                            'label': _get_asset_header_label(ASSET_MANAGEMENT_DEPARTMENT_COLUMN),
                            'before': before_department,
                            'after': after_department,
                        },
                        {
                            'column': ASSET_MANAGEMENT_ASSET_NO_COLUMN,
                            'label': _get_asset_header_label(ASSET_MANAGEMENT_ASSET_NO_COLUMN),
                            'before': _normalize_asset_column_value(previous_row.get(ASSET_MANAGEMENT_ASSET_NO_COLUMN, '')),
                            'after': moved_asset_no,
                        },
                    ]
                    change_entries.append(
                        {
                            'action_type': 'move_out',
                            'title': _get_asset_history_title(previous_row, len(move_details)),
                            'details': move_details,
                        }
                    )
                    change_entries.append(
                        {
                            'action_type': 'move_in',
                            'title': _get_asset_history_title(row, len(move_details)),
                            'details': move_details,
                        }
                    )
                    previous_identity = _get_asset_row_identity(previous_row)
                    if any(previous_identity):
                        matched_before_keys.add(previous_identity)
                    if isinstance(original_row_index, int):
                        matched_before_indexes.add(original_row_index)
                    # preserve original row so it remains visible after move
                    moved_out_rows.append(dict(previous_row))
                    continue

            if previous_row is None:
                details = _build_asset_added_details(row, headers)
                change_entries.append(
                    {
                        'action_type': 'add',
                        'title': _get_asset_history_title(row, len(details)),
                        'details': details,
                    }
                )
                continue

            if isinstance(original_row_index, int) and original_row_index not in matched_before_indexes:
                matched_before_indexes.add(original_row_index)

            if any(identity):
                matched_before_keys.add(identity)

            details = _build_asset_change_details(previous_row, row, headers)
            if not details:
                continue
            change_entries.append(
                {
                    'action_type': 'edit',
                    'title': _get_asset_history_title(row, len(details)),
                    'details': details,
                }
            )

        # Only treat unmatched before_rows as deletions when we're doing a full replace (no selected_department).
        if not selected_department:
            for before_row in before_rows:
                before_identity = _get_asset_row_identity(before_row)
                before_original_row_index = before_row.get('__original_row_index')
                if not any(before_identity) or before_identity in matched_before_keys:
                    continue
                if isinstance(before_original_row_index, int) and before_original_row_index in matched_before_indexes:
                    continue
                details = _build_asset_deleted_details(before_row, headers)
                change_entries.append(
                    {
                        'action_type': 'delete',
                        'title': _get_asset_history_title(before_row, len(details)),
                        'details': details,
                    }
                )

        # Ensure moved-out originals are included in the payload to preserve old records
        if moved_out_rows:
            existing_identities = { _get_asset_row_identity(r) for r in merged_payload if any(_get_asset_row_identity(r)) }
            for mor in moved_out_rows:
                ident = _get_asset_row_identity(mor)
                if not any(ident) or ident in existing_identities:
                    continue
                merged_payload.append(mor)

            # re-filter out any fully blank rows after adding moved-out originals
            merged_payload = [r for r in merged_payload if not _is_asset_blank_row(r, headers)]

        if not change_entries:
            return jsonify({'success': True, 'message': '변경 사항이 없습니다.'})

        placeholders = ','.join(['%s'] * len(import_ids))
        cursor.execute(
            f"DELETE FROM {ASSET_MANAGEMENT_ROW_TABLE} WHERE import_id IN ({placeholders})",
            tuple(import_ids),
        )
        cursor.execute(
            f"DELETE FROM {ASSET_MANAGEMENT_IMPORT_TABLE} WHERE id IN ({placeholders})",
            tuple(import_ids),
        )
        cursor.execute(
            f"""
            INSERT INTO {ASSET_MANAGEMENT_IMPORT_TABLE}
                (asset_year, source_filename, sheet_name, imported_by, headers_json, row_count)
            VALUES
                (%s, %s, %s, %s, %s, %s)
            """,
            (
                year,
                'manual_edit',
                'manual_edit',
                (session.get('user', {}) or {}).get('name', ''),
                json.dumps(headers, ensure_ascii=False),
                len(merged_payload),
            ),
        )
        import_id = cursor.lastrowid
        cursor.executemany(
            f"""
            INSERT INTO {ASSET_MANAGEMENT_ROW_TABLE}
                (import_id, row_index, row_data)
            VALUES
                (%s, %s, %s)
            """,
            [
                (import_id, index + 1, json.dumps(row_data, ensure_ascii=False))
                for index, row_data in enumerate(merged_payload)
            ],
        )
        for change_entry in change_entries:
            _record_asset_management_history(
                cursor,
                year,
                change_entry['action_type'],
                change_entry['title'],
                change_entry['details'],
            )
        db.commit()
        return jsonify({'success': True, 'message': f'{len(change_entries)}건의 변경사항을 저장했습니다.'})
    except mysql.connector.Error as e:
        if db:
            db.rollback()
        print(f'[ERROR] 재물관리 수정 저장 실패: {e}')
        return jsonify({'success': False, 'message': '수정 저장 중 오류가 발생했습니다.'}), 500
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()


@bp.route('/api/asset-management/export', methods=['POST'])
def export_asset_management_excel():
    if not _has_admin_access():
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    payload = request.get_json(silent=True) or {}
    year = payload.get('year')

    try:
        year = int(year)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': '연도 정보가 올바르지 않습니다.'}), 400

    payload_rows = payload.get('rows')
    rows: list[dict[str, object]]
    headers = list(ASSET_MANAGEMENT_FIXED_HEADERS)
    show_modified_only = _is_asset_modified_only_enabled(payload.get('show_modified'))

    if isinstance(payload_rows, list):
        rows = [row for row in payload_rows if isinstance(row, dict)]
    else:
        selected_filters = _normalize_asset_selected_filters(payload.get('filters') if isinstance(payload.get('filters'), dict) else {})
        db = create_connection()
        if db is None:
            return jsonify({'success': False, 'message': 'Database connection could not be established'}), 500

        cursor = None
        try:
            cursor = db.cursor(dictionary=True)
            _, headers, rows = _load_asset_management_view_data(cursor, year)
            rows = _apply_asset_management_filters(rows, selected_filters)
            if show_modified_only:
                history_rows = _load_asset_management_history(cursor, year)
                highlighted_row_keys = {_extract_asset_history_key(item.get('change_summary', '')) for item in history_rows if item.get('change_summary')}
                rows = [row for row in rows if _get_asset_row_highlight_key(row) in highlighted_row_keys]
            rows = _sort_asset_management_rows(rows, selected_filters)
        except mysql.connector.Error as e:
            print(f'[ERROR] 재물관리 엑셀 변환 조회 실패: {e}')
            return jsonify({'success': False, 'message': '엑셀 변환용 데이터를 불러오지 못했습니다.'}), 500
        finally:
            if cursor:
                cursor.close()
            if db:
                db.close()

    output = _build_asset_export_workbook(year, headers, rows)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'{year}_재물관리.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/api/project_progress/<contract_code>', methods=['GET'])
def api_project_progress(contract_code: str):
    try:
        progress = calc_progress(None, contract_code)
        return jsonify({'contract_code': contract_code, 'progress': round(float(progress), 2)})
    except Exception as e:
        return jsonify({'error': True, 'message': str(e)}), 500


@bp.route('/api/check_contract_code/<contract_code>', methods=['GET'])
def check_contract_code(contract_code: str):
    try:
        connection = create_connection()
        if connection is None:
            return jsonify({'error': True, 'message': '데이터베이스 연결에 실패했습니다.'}), 500

        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT COUNT(*) 
            FROM Projects 
            WHERE ContractCode = %s
            """,
            (contract_code,),
        )
        count = cursor.fetchone()[0]
        cursor.close()
        connection.close()

        return jsonify({
            'exists': count > 0,
            'message': '이미 존재하는 계약코드입니다.' if count > 0 else '사용 가능한 계약코드입니다.',
        })

    except mysql.connector.Error as e:
        print(f"Database error: {str(e)}")
        return jsonify({'error': True, 'message': '데이터베이스 조회 중 오류가 발생했습니다.'}), 500

    except Exception as e:
        print(f"Error checking contract code: {str(e)}")
        return jsonify({'error': True, 'message': '계약코드 확인 중 오류가 발생했습니다.'}), 500


@bp.route('/api/search_projects', methods=['GET'])
def search_projects():
    try:
        search_term = request.args.get('term', '').strip()
        year = request.args.get('year', type=int)
        search_type = request.args.get('type', '')

        if not search_term:
            return jsonify({"projects": [], "current_page": 1, "total_pages": 1})

        connection = create_connection()
        if connection is None:
            return jsonify({"error": "Database connection failed"}), 500

        cursor = connection.cursor(dictionary=True)

        base_query = """
            FROM Projects 
            WHERE (ContractCode LIKE %s OR ProjectName LIKE %s OR orderPlace LIKE %s)
        """
        params = [f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"]

        if year:
            base_query += " AND YEAR(StartDate) = %s AND ContractCode NOT LIKE '%%검토%%'"
            params.append(year)

        if search_type == "yearly":
            base_query += " AND yearProject = 1"
        elif search_type == "examine":
            base_query += " AND ContractCode LIKE '%%검토%%'"

        query = f"""
            SELECT 
                ProjectID, 
                ContractCode, 
                ProjectName, 
                orderPlace,
                yearProject,
                project_status,
                outsourcingCheck
            {base_query}
            ORDER BY ContractCode ASC 
            LIMIT %s OFFSET %s
        """

        page = request.args.get("page", 1, type=int)
        per_page = 20
        offset = (page - 1) * per_page
        params.extend([per_page, offset])

        cursor.execute(query, params)
        results = cursor.fetchall()

        contract_codes = [
            (proj.get('ContractCode') or proj.get('contractCode') or proj.get('contract_code'))
            for proj in results if isinstance(proj, dict)
        ]
        progress_map = {}
        if contract_codes:
            try:
                progress_map = calc_progress_bulk(contract_codes) or {}
            except Exception:
                progress_map = {}
        for proj in results:
            code = None
            if isinstance(proj, dict):
                code = proj.get('ContractCode') or proj.get('contractCode') or proj.get('contract_code')
            proj['progress'] = float(progress_map.get(code, 0.0)) if code else 0.0

        risk_map = {}
        contract_codes_distinct = [c for c in set(contract_codes) if c]
        if contract_codes_distinct:
            try:
                has_division = False
                try:
                    cursor.execute("SHOW COLUMNS FROM project_risks LIKE 'division'")
                    has_division = cursor.fetchone() is not None
                except Exception:
                    has_division = False

                fmt = ','.join(['%s'] * len(contract_codes_distinct))
                if has_division:
                    cursor.execute(
                        f"""
                        SELECT DISTINCT contractcode FROM project_risks
                        WHERE contractcode IN ({fmt})
                          AND (division IS NULL OR division <> '완료')
                        """,
                        tuple(contract_codes_distinct),
                    )
                else:
                    cursor.execute(
                        f"""
                        SELECT DISTINCT contractcode FROM project_risks
                        WHERE contractcode IN ({fmt})
                        """,
                        tuple(contract_codes_distinct),
                    )
                for r in cursor.fetchall() or []:
                    code = r.get('contractcode') if isinstance(r, dict) else None
                    if code:
                        risk_map[code] = True
            except Exception:
                risk_map = {}
        for proj in results:
            if isinstance(proj, dict):
                code = proj.get('ContractCode') or proj.get('contractCode') or proj.get('contract_code')
                proj['has_risk'] = bool(risk_map.get(code))

        count_query = f"SELECT COUNT(*) as count {base_query}"
        cursor.execute(count_query, params[:-2])
        total_projects = cursor.fetchone()["count"]
        total_pages = max((total_projects + per_page - 1) // per_page, 1)

        cursor.close()
        connection.close()

        return jsonify({"projects": results, "current_page": page, "total_pages": total_pages})

    except Exception as e:
        print(f"Error searching projects: {str(e)}")
        return jsonify({"error": "검색 중 오류가 발생했습니다."}), 500


@bp.route('/save_staff', methods=['POST'])
def save_staff():
    data = request.get_json()
    db = create_connection()

    if db is None:
        return jsonify({"success": False, "message": "DB 연결 실패"}), 500

    cursor = None
    try:
        cursor = db.cursor(dictionary=True)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        base_password = "1q2w3e4r!"
        default_password = hashlib.sha256(base_password.encode()).hexdigest()

        cursor.execute(
            """
            SELECT userID, Name, Password, Department, Position, EmpNo, JoinDate, Phone, Auth, note, CreateDate,
                   COALESCE(adminAUTH, 0)  AS adminAUTH,
                   COALESCE(dataauth, 0)   AS dataauth,
                   COALESCE(reportAUTH, 0) AS reportAUTH,
                   COALESCE(projectAUTH, 0) AS projectAUTH,
                   COALESCE(meetingAuth, 0) AS meetingAuth
            FROM users
            """
        )
        existing_users = {
            (row['userID'], row['Name']): {
                'Password': row['Password'],
                'Department': row.get('Department') or '',
                'Position': row.get('Position') or '',
                'EmpNo': row.get('EmpNo') or '',
                'JoinDate': row.get('JoinDate'),
                'Phone': row.get('Phone') or '',
                'Auth': row.get('Auth') or '',
                'note': row.get('note') or '',
                'CreateDate': row.get('CreateDate'),
                'adminAUTH': int(row.get('adminAUTH', 0) or 0),
                'dataauth': int(row.get('dataauth', 0) or 0),
                'reportAUTH': int(row.get('reportAUTH', 0) or 0),
                'projectAUTH': int(row.get('projectAUTH', 0) or 0),
                'meetingAuth': int(row.get('meetingAuth', 0) or 0),
            }
            for row in cursor.fetchall()
        }

        cursor.execute("DELETE FROM users")
        current_session_user = session.get('user') or {}
        current_session_user_id = str(current_session_user.get('userID') or '').strip().lower()
        refreshed_session_user = None

        for user in data:
            userID = user.get('userID')
            name = user.get('Name')
            if not userID or not name:
                continue

            prev = existing_users.get((userID, name))
            password = prev['Password'] if prev else default_password
            department = user.get('Department', prev['Department'] if prev else '')
            auth = user.get('Auth', prev['Auth'] if prev else '')
            note = user.get('note', prev['note'] if prev else '')
            emp_no = user.get('EmpNo', prev['EmpNo'] if prev else '')
            position = user.get('Position', prev['Position'] if prev else '')
            join_date = user.get('JoinDate', prev['JoinDate'] if prev else None) or None
            phone = user.get('Phone', prev['Phone'] if prev else '')
            create_date = prev['CreateDate'] if prev and prev.get('CreateDate') else now

            incoming_admin = user.get('adminAUTH', None)
            if incoming_admin is None:
                adminAUTH = prev['adminAUTH'] if prev is not None else 0
            else:
                try:
                    if isinstance(incoming_admin, bool):
                        adminAUTH = 1 if incoming_admin else 0
                    elif isinstance(incoming_admin, (int, float)):
                        adminAUTH = 1 if int(incoming_admin) == 1 else 0
                    elif isinstance(incoming_admin, str):
                        adminAUTH = 1 if incoming_admin.lower() in ('1', 'true', 'y', 'yes') else 0
                    else:
                        adminAUTH = 0
                except Exception:
                    adminAUTH = 0

            incoming = user.get('dataauth', None)
            if incoming is None:
                dataauth = prev['dataauth'] if prev is not None else 0
            else:
                try:
                    if isinstance(incoming, bool):
                        dataauth = 1 if incoming else 0
                    elif isinstance(incoming, (int, float)):
                        dataauth = 1 if int(incoming) == 1 else 0
                    elif isinstance(incoming, str):
                        dataauth = 1 if incoming.lower() in ('1', 'true', 'y', 'yes') else 0
                    else:
                        dataauth = 0
                except Exception:
                    dataauth = 0

            incoming_report = user.get('reportAUTH', None)
            if incoming_report is None:
                reportAUTH = prev['reportAUTH'] if prev is not None else 0
            else:
                try:
                    if isinstance(incoming_report, bool):
                        reportAUTH = 1 if incoming_report else 0
                    elif isinstance(incoming_report, (int, float)):
                        reportAUTH = 1 if int(incoming_report) == 1 else 0
                    elif isinstance(incoming_report, str):
                        reportAUTH = 1 if incoming_report.lower() in ('1', 'true', 'y', 'yes') else 0
                    else:
                        reportAUTH = 0
                except Exception:
                    reportAUTH = 0

            incoming_project = user.get('projectAUTH', None)
            if incoming_project is None:
                projectAUTH = prev['projectAUTH'] if prev is not None else 0
            else:
                try:
                    if isinstance(incoming_project, bool):
                        projectAUTH = 1 if incoming_project else 0
                    elif isinstance(incoming_project, (int, float)):
                        projectAUTH = 1 if int(incoming_project) == 1 else 0
                    elif isinstance(incoming_project, str):
                        projectAUTH = 1 if incoming_project.lower() in ('1', 'true', 'y', 'yes') else 0
                    else:
                        projectAUTH = 0
                except Exception:
                    projectAUTH = 0

            incoming_meeting = user.get('meetingAuth', None)
            if incoming_meeting is None:
                meetingAuth = prev['meetingAuth'] if prev is not None else 0
            else:
                try:
                    if isinstance(incoming_meeting, bool):
                        meetingAuth = 1 if incoming_meeting else 0
                    elif isinstance(incoming_meeting, (int, float)):
                        meetingAuth = 1 if int(incoming_meeting) == 1 else 0
                    elif isinstance(incoming_meeting, str):
                        meetingAuth = 1 if incoming_meeting.lower() in ('1', 'true', 'y', 'yes') else 0
                    else:
                        meetingAuth = 0
                except Exception:
                    meetingAuth = 0

            cursor.execute(
                """
                INSERT INTO users (userID, Password, Name, Department, Position, EmpNo, JoinDate, Phone, Auth, note, adminAUTH, dataauth, reportAUTH, projectAUTH, meetingAuth, CreateDate, UpdateDate)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    userID,
                    password,
                    name,
                    department,
                    position,
                    emp_no,
                    join_date,
                    phone,
                    auth,
                    note,
                    adminAUTH,
                    dataauth,
                    reportAUTH,
                    projectAUTH,
                    meetingAuth,
                    create_date,
                    now,
                ),
            )

            if current_session_user_id and str(userID).strip().lower() == current_session_user_id:
                refreshed_session_user = {
                    **current_session_user,
                    'userID': userID,
                    'name': name,
                    'Name': name,
                    'department': department,
                    'Department': department,
                    'position': position,
                    'Position': position,
                    'auth': auth,
                    'Auth': auth,
                    'adminAUTH': adminAUTH,
                    'adminauth': adminAUTH,
                    'dataauth': dataauth,
                    'reportAUTH': reportAUTH,
                    'projectAUTH': projectAUTH,
                    'meetingAuth': meetingAuth,
                    'meetingauth': meetingAuth,
                }

        db.commit()
        if refreshed_session_user is not None:
            session['user'] = refreshed_session_user
        return jsonify({"success": True})

    except mysql.connector.Error as e:
        print("[DB ERROR]", e)
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        db.close()
