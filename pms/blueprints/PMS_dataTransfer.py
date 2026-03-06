from __future__ import annotations

from datetime import date, datetime
import re

import mysql.connector
from flask import Blueprint, jsonify, render_template, request
from openpyxl import load_workbook
from io import BytesIO

from ..db import create_connection

bp = Blueprint('data_transfer', __name__)


@bp.route('/api/get_transfer_data/')
def get_transfer_data():
    """transferData 테이블에서 사업 목록을 가져오는 API (페이지네이션 포함)"""
    db = create_connection()
    if db is None:
        return jsonify({'error': 'Database connection could not be established'}), 500

    try:
        cursor = db.cursor(dictionary=True)
        page = request.args.get('page', 1, type=int)
        per_page = 20

        cursor.execute('SELECT COUNT(*) AS count FROM transferData')
        total_projects = cursor.fetchone()['count']
        total_pages = max(1, (total_projects + per_page - 1) // per_page)

        if page > total_pages or page < 1:
            page = 1

        offset = (page - 1) * per_page

        cursor.execute(
            """
            SELECT ContractCode, ProjectName
            FROM transferData
            ORDER BY ContractCode DESC
            LIMIT %s OFFSET %s
            """,
            (per_page, offset),
        )
        projects = cursor.fetchall()

    except mysql.connector.Error as e:
        print(f'Error executing SQL query: {e}')
        return jsonify({'error': 'Failed to fetch data'}), 500
    finally:
        try:
            cursor.close()
            db.close()
        except Exception:
            pass

    return jsonify({'projects': projects, 'total_pages': total_pages, 'current_page': page})


@bp.route('/dataTransfer')
def data_transfer():
    return render_template('PMS_dataTransfer.html')


@bp.route('/testmove')
@bp.route('/doc_editor')
def testmove():
    return render_template('doc_editor.html')


def normalize_contractcode(value) -> str:
    if value is None:
        return ''
    contractcode = str(value).strip()
    if not contractcode:
        return ''
    return re.sub(r'\(L\)', '-00', contractcode)


def _extract_engineer_rows_from_excel(file_stream):
    wb = load_workbook(file_stream, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = []
    for cell in ws[1]:
        value = '' if cell.value is None else str(cell.value).strip()
        headers.append(value)

    contract_idx = None
    for idx, header in enumerate(headers):
        if header == '계약번호' or '계약번호' in header:
            contract_idx = idx
            break

    if contract_idx is None:
        raise ValueError('엑셀에서 계약번호 컬럼을 찾지 못했습니다.')

    engineer_columns = []
    for idx, header in enumerate(headers):
        if any(key in header for key in ('사책', '분책', '분참')):
            work_position = re.sub(r'\(.*?\)', '', header).strip()
            if work_position:
                engineer_columns.append((idx, work_position))

    if not engineer_columns:
        raise ValueError('엑셀에서 담당업무 컬럼(사책/분책/분참)을 찾지 못했습니다.')

    extracted_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        raw_contractcode = row[contract_idx] if contract_idx < len(row) else None
        contractcode = normalize_contractcode(raw_contractcode)

        if not contractcode:
            continue
        if '소계' in contractcode:
            continue

        for col_idx, work_position in engineer_columns:
            if col_idx >= len(row):
                continue
            raw_name = row[col_idx]
            if raw_name is None:
                continue
            name = str(raw_name).strip()
            if not name or name == '-':
                continue

            extracted_rows.append((contractcode, work_position, name))

    return extracted_rows


def extract_yellow_rows(file_stream):
    try:
        wb = load_workbook(file_stream, data_only=True)
        sheet_name = '★전체실적정리★'

        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]

        target_columns = {'B': 2, 'D': 4, 'G': 7, 'H': 8, 'I': 9, 'J': 10, 'L': 12}

        extracted_data = []

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            b_cell = row[target_columns['B'] - 1]
            color_code = b_cell.fill.start_color.index if b_cell.fill.start_color is not None else 'No Color'

            is_yellow = color_code in ['FFFF00', 'FFFFFF00']
            if is_yellow:
                row_data = {col: row[idx - 1].value for col, idx in target_columns.items()}
                extracted_data.append(row_data)

        return extracted_data
    except Exception as e:
        print(f'엑셀 데이터 처리 중 오류 발생: {e}')
        return []


@bp.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            print('파일이 업로드되지 않음')
            return jsonify({'error': '파일이 없습니다.'}), 400

        file = request.files['file']
        if file.filename == '':
            print('파일명이 비어 있음')
            return jsonify({'error': '파일이 선택되지 않았습니다.'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            print('잘못된 파일 형식 업로드됨')
            return jsonify({'error': '엑셀 파일만 업로드 가능합니다.'}), 400

        print(f'업로드된 파일: {file.filename}')
        file_stream = BytesIO(file.read())
        extracted_data = extract_yellow_rows(file_stream)

        for row in extracted_data:
            row['D'] = convert_date_format(row['D'])
            row['G'] = convert_date_format(row['G'])

        return jsonify(extracted_data)
    except Exception as e:
        print(f'파일 업로드 처리 중 오류 발생: {e}')
        return jsonify({'error': str(e)}), 500


@bp.route('/api/engineer_transfer/extract', methods=['POST'])
def extract_engineer_transfer_data():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400

    file = request.files['file']
    if not file or not file.filename:
        return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다.'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '엑셀 파일만 업로드 가능합니다.'}), 400

    try:
        extracted_rows = _extract_engineer_rows_from_excel(BytesIO(file.read()))
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        print(f'[ERROR] 참여기술자 엑셀 추출 실패: {e}')
        return jsonify({'success': False, 'message': '엑셀 처리 중 오류가 발생했습니다.'}), 500

    db = create_connection()
    if db is None:
        return jsonify({'success': False, 'message': 'Database connection could not be established'}), 500

    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            """
            SELECT contractcode, projectname
            FROM projects
            WHERE contractcode IS NOT NULL
              AND contractcode NOT LIKE %s
            """,
            ('%검토%',),
        )
        project_rows = cursor.fetchall() or []
        project_map = {
            normalize_contractcode(row['contractcode']): row.get('projectname') or ''
            for row in project_rows
            if row.get('contractcode')
        }

        cursor.execute(
            """
            SELECT DISTINCT contractcode
            FROM project_engineers
            WHERE contractcode IS NOT NULL
            """
        )
        existing_rows = cursor.fetchall() or []
        existing_contractcodes = {
            normalize_contractcode(row['contractcode'])
            for row in existing_rows
            if row.get('contractcode')
        }
    except Exception as e:
        print(f'[ERROR] 참여기술자 DB 조회 실패: {e}')
        return jsonify({'success': False, 'message': 'DB 조회 중 오류가 발생했습니다.'}), 500
    finally:
        cursor.close()
        db.close()

    eligible_contractcodes = set(project_map.keys()) - existing_contractcodes
    grouped = {}

    for contractcode, work_position, name in extracted_rows:
        contractcode = normalize_contractcode(contractcode)
        if contractcode not in eligible_contractcodes:
            continue

        if contractcode not in grouped:
            grouped[contractcode] = {
                'contractcode': contractcode,
                'project_name': project_map.get(contractcode, ''),
                'rows': [],
                '_dedup': set(),
            }

        key = (work_position, name)
        if key in grouped[contractcode]['_dedup']:
            continue

        grouped[contractcode]['rows'].append({'work_position': work_position, 'name': name})
        grouped[contractcode]['_dedup'].add(key)

    items = []
    for contractcode in sorted(grouped.keys()):
        item = grouped[contractcode]
        item.pop('_dedup', None)
        items.append(item)

    return jsonify(
        {
            'success': True,
            'items': items,
            'total_contractcodes': len(items),
            'total_rows': sum(len(item['rows']) for item in items),
        }
    )


@bp.route('/api/engineer_transfer/insert', methods=['POST'])
def insert_engineer_transfer_data():
    payload = request.get_json(silent=True) or {}
    items = payload.get('items') or []

    if not isinstance(items, list) or not items:
        return jsonify({'success': False, 'message': '삽입할 데이터가 없습니다.'}), 400

    requested_contractcodes = {
        normalize_contractcode(item.get('contractcode'))
        for item in items
        if normalize_contractcode(item.get('contractcode'))
    }
    if not requested_contractcodes:
        return jsonify({'success': False, 'message': '유효한 계약번호가 없습니다.'}), 400

    db = create_connection()
    if db is None:
        return jsonify({'success': False, 'message': 'Database connection could not be established'}), 500

    cursor = db.cursor(dictionary=True)

    try:
        placeholders = ','.join(['%s'] * len(requested_contractcodes))
        params = list(requested_contractcodes)

        cursor.execute(
            f"""
            SELECT contractcode
            FROM projects
            WHERE contractcode IN ({placeholders})
              AND contractcode NOT LIKE %s
            """,
            (*params, '%검토%'),
        )
        valid_rows = cursor.fetchall() or []
        valid_contractcodes = {
            normalize_contractcode(row['contractcode'])
            for row in valid_rows
            if row.get('contractcode')
        }

        cursor.execute(
            f"""
            SELECT DISTINCT contractcode
            FROM project_engineers
            WHERE contractcode IN ({placeholders})
            """,
            tuple(params),
        )
        existing_rows = cursor.fetchall() or []
        existing_contractcodes = {
            normalize_contractcode(row['contractcode'])
            for row in existing_rows
            if row.get('contractcode')
        }

        insert_rows = []
        inserted_contractcodes = set()

        for item in items:
            contractcode = normalize_contractcode(item.get('contractcode'))
            if not contractcode:
                continue
            if contractcode not in valid_contractcodes:
                continue
            if contractcode in existing_contractcodes:
                continue

            for row in item.get('rows') or []:
                work_position = str(row.get('work_position') or '').strip()
                name = str(row.get('name') or '').strip()
                if not work_position or not name:
                    continue
                insert_rows.append((contractcode, work_position, name))
                inserted_contractcodes.add(contractcode)

        if not insert_rows:
            return jsonify({'success': False, 'message': '삽입 가능한 데이터가 없습니다.'}), 400

        cursor.executemany(
            """
            INSERT INTO project_engineers (contractcode, work_position, name)
            VALUES (%s, %s, %s)
            """,
            insert_rows,
        )
        db.commit()

        return jsonify(
            {
                'success': True,
                'message': f"참여기술자 데이터 삽입 완료: {len(insert_rows)}건",
                'inserted_rows': len(insert_rows),
                'inserted_contractcodes': len(inserted_contractcodes),
            }
        )
    except Exception as e:
        db.rollback()
        print(f'[ERROR] 참여기술자 데이터 삽입 실패: {e}')
        return jsonify({'success': False, 'message': 'DB 삽입 중 오류가 발생했습니다.'}), 500
    finally:
        cursor.close()
        db.close()


@bp.route('/saveExcelData', methods=['POST'])
def save_data():
    try:
        extracted_data = request.json
        db = create_connection()
        cursor = db.cursor()

        for row in extracted_data:
            cursor.execute('SELECT COUNT(*) FROM transferData WHERE ContractCode = %s', (row['B'],))
            exists = cursor.fetchone()[0] > 0
            projectCost_NoVAT = round(float(row['J']) / 1.1) if row['J'] else None

            if exists:
                cursor.execute(
                    """
                    UPDATE transferData
                    SET ProjectName = %s, ProjectCost = %s, ProjectCost_NoVAT = %s, StartDate = %s,
                        EndDate = %s, OrderPlace = %s, ContributionRate = %s
                    WHERE ContractCode = %s
                    """,
                    (row['H'], row['J'], projectCost_NoVAT, row['D'], row['G'], row['I'], row['L'], row['B']),
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO transferData (ContractCode, ProjectName, ProjectCost, ProjectCost_NoVAT, StartDate,
                                             EndDate, OrderPlace, ContributionRate, ProjectDetails)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NULL)
                    """,
                    (row['B'], row['H'], row['J'], projectCost_NoVAT, row['D'], row['G'], row['I'], row['L']),
                )

        db.commit()
        return jsonify({'message': '데이터가 성공적으로 저장되었습니다.'})

    except Exception as e:
        print('ERROR:', str(e))
        return jsonify({'message': str(e)}), 500

    finally:
        try:
            cursor.close()
            db.close()
        except Exception:
            pass


@bp.route('/api/get_transfer_detail', methods=['GET'])
def get_transfer_detail():
    contract_code = request.args.get('contractCode')
    if not contract_code:
        return jsonify({'error': 'ContractCode is required'}), 400

    db = create_connection()
    cursor = db.cursor(dictionary=True)

    try:
        cursor.execute(
            """
            SELECT ContractCode, ProjectName, ProjectCost, StartDate, EndDate,ContributionRate,
                   OrderPlace, Manager, ProjectDetails
            FROM transferdata
            WHERE ContractCode = %s
            """,
            (contract_code,),
        )
        data = cursor.fetchone()
        if not data:
            return jsonify({'error': 'No data found'}), 404

        if isinstance(data.get('StartDate'), date):
            data['StartDate'] = data['StartDate'].isoformat()
        if isinstance(data.get('EndDate'), date):
            data['EndDate'] = data['EndDate'].isoformat()

        return jsonify(data)

    except mysql.connector.Error as e:
        return jsonify({'error': f'Database error: {str(e)}'}), 500

    finally:
        cursor.close()
        db.close()


def convert_date_format(date_value):
    """날짜 형식을 'YYYY-MM-DD'로 변환"""
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')
    try:
        return datetime.strptime(date_value, '%a, %d %b %Y %H:%M:%S GMT').strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        return None


@bp.route('/insertProject', methods=['POST'])
def insert_project():
    try:
        extracted_data = request.json
        db = create_connection()
        cursor = db.cursor()

        for row in extracted_data:
            projectCost = float(row['J'].replace(',', '')) if row['J'] else None
            projectCost_NoVAT = round(projectCost / 1.1) if projectCost else None
            changeProjectCost = projectCost

            cursor.execute(
                """
                INSERT INTO projects (ContractCode, ProjectName, ProjectCost, ProjectCost_NoVAT, StartDate,
                                      EndDate, orderPlace, Manager, ContributionRate, ProjectDetails,
                                      AcademicResearchRate, OperationalRate, EquipmentRate, safetyRate,
                                      ChangeProjectCost, outsourcingCheck, yearProject, LinkProjectCheck,
                                      referenceProject1, referenceProject2, referenceProject3, referenceProject4, referenceProject5)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    row['B'],
                    row['H'],
                    projectCost,
                    projectCost_NoVAT,
                    row['D'],
                    row['G'],
                    row['I'],
                    row['manager'],
                    row['L'],
                    row['projectDetails'],
                    5.30,
                    25.053,
                    2.30,
                    None,
                    changeProjectCost,
                    0,
                    0,
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                ),
            )

            cursor.execute('DELETE FROM transferData WHERE ContractCode = %s', (row['B'],))

        db.commit()
        return jsonify({'message': '데이터가 성공적으로 저장되었고, 기존 데이터가 삭제되었습니다.'})

    except Exception as e:
        print('ERROR:', str(e))
        return jsonify({'message': str(e)}), 500

    finally:
        cursor.close()
        db.close()
